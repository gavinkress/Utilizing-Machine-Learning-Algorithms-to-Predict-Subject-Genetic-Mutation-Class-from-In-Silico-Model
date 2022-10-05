import numpy as np
import openpyxl
import pandas as pd
import pyexcel as p
from itertools import groupby
import time
import nitime
path = '~\PycharmProjects\Variable_Analysis\Models'
path2 = 'C:/Users/gavin/PycharmProjects\Variable_Analysis\Models'
strArr = ['\Disease', '\Healthy', '\DiseaseLOF']

for ww in strArr:
    for kk in range(1, 76):
        start_time = time.time()
        model = ww + str(kk)
        print("---  Starting " + model + "---" )
        print("--- (1/14) Starting to Extract Embedding Dimension Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60) , (time.time() - start_time) % 60))
        start_time_ed = time.time()
        nn = 200  # Number of Neurons

        # Open Data
        vdf = pd.read_excel(path + model + '\Voltage.xlsx')

        vdf2 = vdf.rename(columns={"Unnamed: 0": "Time"})
        vdf2['Time'] = pd.to_timedelta(vdf2['Time'].div(10), 'ms')
        vdf2 = vdf2.set_index(vdf2['Time']).resample('1500us').max().reset_index(drop=True)
        vdf2['Time'] = pd.to_numeric(vdf2['Time'])
        v = pd.DataFrame.to_numpy(vdf2)
        t = v[:, 1]

        mea = pd.read_excel(path + model + '\MEA.xls', header=None).to_numpy()


        # FNN Function

        def fnn(x, tau, m, rtol, atol):
            L = len(x)
            Ra = np.std(x)
            FNN = np.zeros(m)
            for d in range(1, m + 1):
                M = L - d * tau - 1
                Y = np.zeros([M, d])

                for i in range(0, d):
                    Y[:, i] = x[i * tau:(M + i * tau)]  # Create M vectors in d dimensions

                for n in range(1, M):

                    y0 = np.ones([M, 1])

                    distance = np.zeros([M, 2])
                    distance[:, 0] = np.sqrt(
                        np.sum(np.power(Y - y0.dot([Y[n, :]]), 2), axis=1))  # get distances of each vector from nth vector
                    distance[:, 1] = range(M)
                    ND = distance[np.argsort(distance[:, 0])]
                    neardis = ND[:, 0]
                    nearpos = ND[:, 1]
                    nearpos = nearpos.astype(int)

                    G = np.abs(x[n + (d) * tau] - x[nearpos[1] + (d) * tau])
                    R = np.sqrt(G ** 2 + neardis[1] ** 2)

                    if G / neardis[2] > rtol or R / Ra > atol:
                        FNN[d - 1] = FNN[d - 1] + 1
            FNN = np.divide(FNN, FNN[0]) * 100
            return FNN


        # Embedding Dimension Test

        def ED(FNN):
            for i in range(np.shape(FNN)[0]):
                EDD = FNN[i]
                if i > 0:
                    if np.abs(FNN[i] - FNN[i - 1]) < 0.2:
                        break
                else:
                    EDD = FNN[i]
            EDD = i
            return EDD


        # Extract Embedding For Each Neuron

        m = 10  # max dimensions to plot
        tau = int(0.05 * len(t))  # time delay
        rtol = 10  # threshold fold distance increase
        atol = 2  # standard deviation threshold
        EDD = np.zeros(np.shape(v)[1] - 2)

        for i in range(2, np.shape(v)[1]):
            EDD[i - 2] = ED(fnn(v[:, i], tau, m, rtol, atol))


        # Embedding Dimension Gradient

        def gradient(A, spac):
            dims = np.shape(A)
            G_x = np.zeros(dims)
            G_y = np.zeros(dims)
            for i in range(dims[0]):
                for j in range(dims[1]):
                    if i == dims[0] - 1 or j == dims[1] - 1:
                        G_x[i, j] = 0
                        G_y[i, j] = 0
                    else:
                        G_x[i, j] = (A[i, j + 1] - A[i, j]) / spac
                        G_y[i, j] = (A[i + 1, j] - A[i, j]) / spac
            return G_x, G_y


        ss = 10
        um = 10 ** (-6)
        data = np.zeros(np.shape(mea))
        for i in range(np.shape(mea)[1] * np.shape(mea)[0]):
            data[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = EDD[i]

        Gr_x, Gr_y = gradient(data, ss)
        Gr = np.power(np.power(Gr_x, 2) + np.power(Gr_y, 2), (1 / 2))

        # Relevant Variables Save


        ED_m = np.mean(EDD)
        ED_max = np.max(EDD)
        ED_std = np.std(EDD)

        ED_gr_m = np.mean(Gr)
        ED_gr_max = np.max(Gr)
        ED_gr_std = np.std(Gr)

        p.save_book_as(file_name=(path + model + '\Variables.xls'),
                       dest_file_name=(path2 + model + '\Variables.xlsx'))

        xfile = openpyxl.load_workbook(path2 + model + '\Variables.xlsx')

        xfile["Sheet 1"]['A5'] = 'MEAN EMBEDDING DIMENSION'
        xfile["Sheet 1"]['A6'] = 'MAX EMBEDDING DIMENSION'
        xfile["Sheet 1"]['A7'] = 'EMBEDDING DIMENSION STD'
        xfile["Sheet 1"]['A8'] = 'MEAN EMBEDDING DIMENSION GRADIENT'
        xfile["Sheet 1"]['A9'] = 'MAX EMBEDDING DIMENSION GRADIENT'
        xfile["Sheet 1"]['A10'] = 'EMBEDDING DIMENSION GRADIENT STD'

        xfile["Sheet 1"]['B5'] = ED_m
        xfile["Sheet 1"]['B6'] = ED_max
        xfile["Sheet 1"]['B7'] = ED_std
        xfile["Sheet 1"]['B8'] = ED_gr_m
        xfile["Sheet 1"]['B9'] = ED_gr_max
        xfile["Sheet 1"]['B10'] = ED_gr_std

        xfile.save(path2 + model + '\Variables.xlsx')

        print("--- (2/14) Extracting Embedding Dimension Variables took %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time_ed)/60) , (time.time() - start_time_ed) % 60))
        print("--- (3/14) Finished Extracting Embedding Dimension Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60) , (time.time() - start_time) % 60))



        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################


        print("--- (4/14) Starting to Extract MultiScale Entropy Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60) , (time.time() - start_time) % 60))
        start_time_mse = time.time()


        m = 2   # dimension of euclidian map
        NS = 140  # Number of Scales


        def MultiScaleEntropy(V, m, r, dd):
            N = len(V)

            # divide time series
            xmi = np.array([V[i: i + m*dd:dd] for i in range(N - m*dd)])
            xmj = np.array([V[i: i + m*dd:dd] for i in range(1, N - m*dd + 1)])

            # Save all matches minus the self-match, compute B
            B = np.sum([np.sum(np.abs(xmii - xmj).max(axis=1) <= r)  for xmii in xmi])


            # Same for A
            m += 1
            xm = np.array([V[i: i + m] for i in range(N - m + 1)])

            A = np.sum([np.sum(np.abs(xmi - xm).max(axis=1) <= r)  for xmi in xm])

            # Return SampEn
            return -np.log(B / A)

        MSE_i = np.zeros(nn)  # integral
        MSE_m = np.zeros(nn)  # mean
        MSE_std = np.zeros(nn)  # std

        for i in range(2,nn):
            r = np.std(v[:, i])*0.5  # max Chebyshev distance
            MSE = [0]*NS
            for dd in range(1, NS+1):
                MSE[dd-1] = MultiScaleEntropy(v[:, i], m, r, dd)

            MSE_i[i-2] = np.sum(MSE)  # integral
            MSE_m[i-2] = np.mean(MSE)  # mean
            MSE_std[i-2] = np.std(MSE)  # std


        data1 = data*0
        data2 = data*0
        data3 = data*0
        for i in range(np.shape(mea)[1] * np.shape(mea)[0]):
            data1[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = MSE_i[i]
            data2[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = MSE_m[i]
            data3[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = MSE_std[i]

        Gr_x1, Gr_y1 = gradient(data1, ss)
        Gr_MSE_i = np.power(np.power(Gr_x1, 2) + np.power(Gr_y1, 2), (1 / 2))
        Gr_x2, Gr_y2 = gradient(data2, ss)
        Gr_MSE_m = np.power(np.power(Gr_x2, 2) + np.power(Gr_y2, 2), (1 / 2))
        Gr_x3, Gr_y3 = gradient(data3, ss)
        Gr_MSE_std = np.power(np.power(Gr_x3, 2) + np.power(Gr_y3, 2), (1 / 2))



        # Relevant Variables Save


        MSE_i_m = np.mean(MSE_i)
        MSE_i_max = np.max(MSE_i)
        MSE_i_std = np.std(MSE_i)
        MSE_i_gr_m = np.mean(Gr_MSE_i)
        MSE_i_gr_max = np.max(Gr_MSE_i)
        MSE_i_gr_std = np.std(Gr_MSE_i)

        MSE_m_m = np.mean(MSE_m)
        MSE_m_max = np.max(MSE_m)
        MSE_m_std = np.std(MSE_m)
        MSE_m_gr_m = np.mean(Gr_MSE_m)
        MSE_m_gr_max = np.max(Gr_MSE_m)
        MSE_m_gr_std = np.std(Gr_MSE_m)

        MSE_std_m = np.mean(MSE_std)
        MSE_std_max = np.max(MSE_std)
        MSE_std_std = np.std(MSE_std)
        MSE_std_gr_m = np.mean(Gr_MSE_std)
        MSE_std_gr_max = np.max(Gr_MSE_std)
        MSE_std_gr_std = np.std(Gr_MSE_std)

        xfile = openpyxl.load_workbook(path2 + model + '\Variables.xlsx')

        xfile["Sheet 1"]['A11'] = 'MEAN MULTISCALE ENTROPY INTEGRAL'
        xfile["Sheet 1"]['A12'] = 'MAX MULTISCALE ENTROPY INTEGRAL'
        xfile["Sheet 1"]['A13'] = 'MULTISCALE ENTROPY INTEGRAL STD'
        xfile["Sheet 1"]['A14'] = 'MEAN MULTISCALE ENTROPY INTEGRAL GRADIENT'
        xfile["Sheet 1"]['A15'] = 'MAX MULTISCALE ENTROPY INTEGRAL GRADIENT'
        xfile["Sheet 1"]['A16'] = 'MULTISCALE ENTROPY INTEGRAL GRADIENT STD'

        xfile["Sheet 1"]['B11'] = MSE_i_m
        xfile["Sheet 1"]['B12'] = MSE_i_max
        xfile["Sheet 1"]['B13'] = MSE_i_std
        xfile["Sheet 1"]['B14'] = MSE_i_gr_m
        xfile["Sheet 1"]['B15'] = MSE_i_gr_max
        xfile["Sheet 1"]['B16'] = MSE_i_gr_std

        xfile["Sheet 1"]['A17'] = 'MEAN MULTISCALE ENTROPY MEAN'
        xfile["Sheet 1"]['A18'] = 'MAX MULTISCALE ENTROPY MEAN'
        xfile["Sheet 1"]['A19'] = 'MULTISCALE ENTROPY MEAN STD'
        xfile["Sheet 1"]['A20'] = 'MEAN MULTISCALE ENTROPY MEAN GRADIENT'
        xfile["Sheet 1"]['A21'] = 'MAX MULTISCALE ENTROPY MEAN GRADIENT'
        xfile["Sheet 1"]['A22'] = 'MULTISCALE ENTROPY MEAN GRADIENT STD'

        xfile["Sheet 1"]['B17'] = MSE_m_m
        xfile["Sheet 1"]['B18'] = MSE_m_max
        xfile["Sheet 1"]['B19'] = MSE_m_std
        xfile["Sheet 1"]['B20'] = MSE_m_gr_m
        xfile["Sheet 1"]['B21'] = MSE_m_gr_max
        xfile["Sheet 1"]['B22'] = MSE_m_gr_std

        xfile["Sheet 1"]['A23'] = 'MEAN MULTISCALE ENTROPY STD'
        xfile["Sheet 1"]['A24'] = 'MAX MULTISCALE ENTROPY STD'
        xfile["Sheet 1"]['A25'] = 'MULTISCALE ENTROPY STD STD'
        xfile["Sheet 1"]['A26'] = 'MEAN MULTISCALE ENTROPY STD GRADIENT'
        xfile["Sheet 1"]['A27'] = 'MAX MULTISCALE ENTROPY STD GRADIENT'
        xfile["Sheet 1"]['A28'] = 'MULTISCALE ENTROPY STD GRADIENT STD'

        xfile["Sheet 1"]['B23'] = MSE_std_m
        xfile["Sheet 1"]['B24'] = MSE_std_max
        xfile["Sheet 1"]['B25'] = MSE_std_std
        xfile["Sheet 1"]['B26'] = MSE_std_gr_m
        xfile["Sheet 1"]['B27'] = MSE_std_gr_max
        xfile["Sheet 1"]['B28'] = MSE_std_gr_std

        xfile.save(path2 + model + '\Variables.xlsx')


        print("--- (5/14) Extracting MultiScale Entropy Variables took %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time_mse)/60) , (time.time() - start_time_mse) % 60))
        print("--- (6/14) Finished Extracting MultiScale Entropy Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60) , (time.time() - start_time) % 60))



        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################

        print("--- (7/14) Starting to Extract Waveform Shape Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60) , (time.time() - start_time) % 60))
        start_time_ws = time.time()
        v = pd.DataFrame.to_numpy(vdf)

        T = 0.95  # Threshold for AP ID (percent of average to cross)
        IIT = 0.5  # interictal threshold (percent of average AP duration)
        PDST = 1.5  # Paroxysmal Depolarization Shift Threshold
        ffi = 200.  # Frequency Interval ms/2000th must be even
        g = v[:, 2:]*0
        gg= g
        NN = np.zeros((nn, 6))
        for n in range(nn):
            gg[:, n]= (v[:, n+2]>(T*np.average(v[:, n+2])))*1.
            g[:, n] = gg[:, n]**(np.max(v[:, n+2])-np.min(v[:, n+2]))+np.min(v[:, n+2])

            dp = np.array([sum(1 for _ in group) for _, group in groupby(g[:, n])])
            x = np.zeros((2, len(dp)))
            y = np.zeros((2, len(dp)))

            if gg[0, n] == 0:
                x[0, :] = dp
                x[1, :] = np.array(list(range(len(dp)))) % 2  # starts with 0

                y[0, :] = dp
                y[1, :] = np.array(list(range(len(dp)))) % 2  # starts with 0
            else:
                x[0, :] = dp
                x[1, :] = (np.array(list(range(len(dp))))+1) % 2  # starts with 1

                y[0, :] = dp
                y[1, :] = (np.array(list(range(len(dp)))) + 1) % 2  # starts with 1

            X = np.multiply(x[0, :], x[1, :])
            X = X[X != 0]

            Y = (X < IIT * np.average(X)) * 1.  # magenta
            X = (X > PDST * np.average(X)) * 1.  # green

            x[1,:]=np.zeros(len(x[1,:]))
            y[1, :] = np.zeros(len(y[1, :]))

            if gg[0, n]==0:
                x[1, (np.array(np.where(X == 1)) + 1) * 2 - 1] = 1
                y[1, (np.array(np.where(Y == 1)) + 1) * 2 - 1] = 1
            else:
                x[1, np.array(np.where(X == 1)) * 2] = 1
                y[1, np.array(np.where(Y == 1)) * 2] = 1


            F = []
            J = []
            for j in range(len(x[0, :])):
                F = np.append(F, np.repeat(int(x[1, j]), x[0, j]))
                J = np.append(J, np.repeat(int(y[1, j]), y[0, j]))

            NN[n, :] = [np.sum(x[1, :]), np.sum(y[1, :]), np.average(F), np.average(J), np.average(x[0, :]*x[1, :]), np.average(y[0, :]*y[1, :])]

        LL = np.zeros(nn)
        Area = np.zeros(nn)

        for n in range(nn):
            LL[n] = np.sum(np.power(np.power(((v[0:-1, 1]-v[1:, 1])*10**-3), 2) + np.power(((v[0:-1, n+2]-v[1:, n+2])*10**-3), 2), 1/2))
            Area[n] = np.sum(v[:, n+2]*10**-4)



        PDS_f = NN[:, 0]
        ISS_f = NN[:, 1]
        PDS_md = NN[:, 4]
        ISS_md = NN[:, 5]


        data1 = data*0
        data2 = data*0
        data3 = data*0
        data4 = data*0
        data5 = data*0
        data6 = data*0
        for i in range(np.shape(mea)[1] * np.shape(mea)[0]):
            data1[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = PDS_f[i]
            data2[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = ISS_f[i]
            data3[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = PDS_md[i]
            data4[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = ISS_md[i]
            data5[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = LL[i]
            data6[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = Area[i]

        Gr_x1, Gr_y1 = gradient(data1, ss)
        Gr_PDS_f = np.power(np.power(Gr_x1, 2) + np.power(Gr_y1, 2), (1 / 2))
        Gr_x2, Gr_y2 = gradient(data2, ss)
        Gr_ISS_f = np.power(np.power(Gr_x2, 2) + np.power(Gr_y2, 2), (1 / 2))
        Gr_x3, Gr_y3 = gradient(data3, ss)
        Gr_PDS_md = np.power(np.power(Gr_x3, 2) + np.power(Gr_y3, 2), (1 / 2))
        Gr_x4, Gr_y4 = gradient(data4, ss)
        Gr_ISS_md = np.power(np.power(Gr_x4, 2) + np.power(Gr_y4, 2), (1 / 2))
        Gr_x5, Gr_y5 = gradient(data5, ss)
        Gr_LL = np.power(np.power(Gr_x5, 2) + np.power(Gr_y5, 2), (1 / 2))
        Gr_x6, Gr_y6 = gradient(data6, ss)
        Gr_Area = np.power(np.power(Gr_x6, 2) + np.power(Gr_y6, 2), (1 / 2))


        PDS_f_m = np.mean(PDS_f)
        PDS_f_max = np.max(PDS_f)
        PDS_f_std = np.std(PDS_f)
        PDS_f_gr_m = np.mean(Gr_PDS_f)
        PDS_f_gr_max = np.max(Gr_PDS_f)
        PDS_f_gr_std = np.std(Gr_PDS_f)

        ISS_f_m = np.mean(ISS_f)
        ISS_f_max = np.max(ISS_f)
        ISS_f_std = np.std(ISS_f)
        ISS_f_gr_m = np.mean(Gr_ISS_f)
        ISS_f_gr_max = np.max(Gr_ISS_f)
        ISS_f_gr_std = np.std(Gr_ISS_f)

        PDS_md_m = np.mean(PDS_md)
        PDS_md_max = np.max(PDS_md)
        PDS_md_std = np.std(PDS_md)
        PDS_md_gr_m = np.mean(Gr_PDS_md)
        PDS_md_gr_max = np.max(Gr_PDS_md)
        PDS_md_gr_std = np.std(Gr_PDS_md)

        ISS_md_m = np.mean(ISS_md)
        ISS_md_max = np.max(ISS_md)
        ISS_md_std = np.std(ISS_md)
        ISS_md_gr_m = np.mean(Gr_ISS_md)
        ISS_md_gr_max = np.max(Gr_ISS_md)
        ISS_md_gr_std = np.std(Gr_ISS_md)

        LL_m = np.mean(LL)
        LL_max = np.max(LL)
        LL_std = np.std(LL)
        LL_gr_m = np.mean(Gr_LL)
        LL_gr_max = np.max(Gr_LL)
        LL_gr_std = np.std(Gr_LL)

        Area_m = np.mean(Area)
        Area_max = np.max(Area)
        Area_std = np.std(Area)
        Area_gr_m = np.mean(Gr_Area)
        Area_gr_max = np.max(Gr_Area)
        Area_gr_std = np.std(Gr_Area)


        xfile = openpyxl.load_workbook(path2 + model + '\Variables.xlsx')

        xfile["Sheet 1"]['A29'] = 'MEAN PAROXYSMAL DEPOLARIZATION SHIFT FREQUENCY'
        xfile["Sheet 1"]['A30'] = 'MAX PAROXYSMAL DEPOLARIZATION SHIFT FREQUENCY'
        xfile["Sheet 1"]['A31'] = 'PAROXYSMAL DEPOLARIZATION SHIFT FREQUENCY STD'
        xfile["Sheet 1"]['A32'] = 'MEAN PAROXYSMAL DEPOLARIZATION SHIFT FREQUENCY GRADIENT'
        xfile["Sheet 1"]['A33'] = 'MAX PAROXYSMAL DEPOLARIZATION SHIFT FREQUENCY GRADIENT'
        xfile["Sheet 1"]['A34'] = 'PAROXYSMAL DEPOLARIZATION SHIFT FREQUENCY GRADIENT STD'

        xfile["Sheet 1"]['B29'] = PDS_f_m
        xfile["Sheet 1"]['B30'] = PDS_f_max
        xfile["Sheet 1"]['B31'] = PDS_f_std
        xfile["Sheet 1"]['B32'] = PDS_f_gr_m
        xfile["Sheet 1"]['B33'] = PDS_f_gr_max
        xfile["Sheet 1"]['B34'] = PDS_f_gr_std

        xfile["Sheet 1"]['A35'] = 'MEAN INTERICTAL SPIKE FREQUENCY'
        xfile["Sheet 1"]['A36'] = 'MAX INTERICTAL SPIKE FREQUENCY'
        xfile["Sheet 1"]['A37'] = 'INTERICTAL SPIKE FREQUENCY STD'
        xfile["Sheet 1"]['A38'] = 'MEAN INTERICTAL SPIKE FREQUENCY GRADIENT'
        xfile["Sheet 1"]['A39'] = 'MAX INTERICTAL SPIKE FREQUENCY GRADIENT'
        xfile["Sheet 1"]['A40'] = 'INTERICTAL SPIKE FREQUENCY GRADIENT STD'

        xfile["Sheet 1"]['B35'] = ISS_f_m
        xfile["Sheet 1"]['B36'] = ISS_f_max
        xfile["Sheet 1"]['B37'] = ISS_f_std
        xfile["Sheet 1"]['B38'] = ISS_f_gr_m
        xfile["Sheet 1"]['B39'] = ISS_f_gr_max
        xfile["Sheet 1"]['B40'] = ISS_f_gr_std

        xfile["Sheet 1"]['A41'] = 'MEAN PAROXYSMAL DEPOLARIZATION SHIFT MEAN DURATION'
        xfile["Sheet 1"]['A42'] = 'MAX PAROXYSMAL DEPOLARIZATION SHIFT MEAN DURATION'
        xfile["Sheet 1"]['A43'] = 'PAROXYSMAL DEPOLARIZATION SHIFT MEAN DURATION STD'
        xfile["Sheet 1"]['A44'] = 'MEAN PAROXYSMAL DEPOLARIZATION SHIFT MEAN DURATION GRADIENT'
        xfile["Sheet 1"]['A45'] = 'MAX PAROXYSMAL DEPOLARIZATION SHIFT MEAN DURATION GRADIENT'
        xfile["Sheet 1"]['A46'] = 'PAROXYSMAL DEPOLARIZATION SHIFT MEAN DURATION GRADIENT STD'

        xfile["Sheet 1"]['B41'] = PDS_md_m
        xfile["Sheet 1"]['B42'] = PDS_md_max
        xfile["Sheet 1"]['B43'] = PDS_md_std
        xfile["Sheet 1"]['B44'] = PDS_md_gr_m
        xfile["Sheet 1"]['B45'] = PDS_md_gr_max
        xfile["Sheet 1"]['B46'] = PDS_md_gr_std

        xfile["Sheet 1"]['A47'] = 'MEAN INTERICTAL SPIKE MEAN DURATION'
        xfile["Sheet 1"]['A48'] = 'MAX INTERICTAL SPIKE MEAN DURATION'
        xfile["Sheet 1"]['A49'] = 'INTERICTAL SPIKE DURATION STD'
        xfile["Sheet 1"]['A50'] = 'MEAN INTERICTAL SPIKE MEAN DURATION GRADIENT'
        xfile["Sheet 1"]['A51'] = 'MAX INTERICTAL SPIKE MEAN DURATION GRADIENT'
        xfile["Sheet 1"]['A52'] = 'INTERICTAL SPIKE MEAN DURATION GRADIENT STD'

        xfile["Sheet 1"]['B47'] = ISS_md_m
        xfile["Sheet 1"]['B48'] = ISS_md_max
        xfile["Sheet 1"]['B49'] = ISS_md_std
        xfile["Sheet 1"]['B50'] = ISS_md_gr_m
        xfile["Sheet 1"]['B51'] = ISS_md_gr_max
        xfile["Sheet 1"]['B52'] = ISS_md_gr_std

        xfile["Sheet 1"]['A53'] = 'MEAN VOLTAGE LINE LENGTH'
        xfile["Sheet 1"]['A54'] = 'MAX VOLTAGE LINE LENGTH'
        xfile["Sheet 1"]['A55'] = 'VOLTAGE LINE LENGTH STD'
        xfile["Sheet 1"]['A56'] = 'MEAN VOLTAGE LINE LENGTH GRADIENT'
        xfile["Sheet 1"]['A57'] = 'MAX VOLTAGE LINE LENGTH GRADIENT'
        xfile["Sheet 1"]['A58'] = 'VOLTAGE LINE LENGTH GRADIENT STD'

        xfile["Sheet 1"]['B53'] = LL_m
        xfile["Sheet 1"]['B54'] = LL_max
        xfile["Sheet 1"]['B55'] = LL_std
        xfile["Sheet 1"]['B56'] = LL_gr_m
        xfile["Sheet 1"]['B57'] = LL_gr_max
        xfile["Sheet 1"]['B58'] = LL_gr_std

        xfile["Sheet 1"]['A59'] = 'MEAN VOLTAGE AREA'
        xfile["Sheet 1"]['A60'] = 'MAX VOLTAGE AREA'
        xfile["Sheet 1"]['A61'] = 'VOLTAGE AREA STD'
        xfile["Sheet 1"]['A62'] = 'MEAN VOLTAGE AREA GRADIENT'
        xfile["Sheet 1"]['A63'] = 'MAX VOLTAGE AREA GRADIENT'
        xfile["Sheet 1"]['A64'] = 'VOLTAGE AREA GRADIENT STD'

        xfile["Sheet 1"]['B59'] = Area_m
        xfile["Sheet 1"]['B60'] = Area_max
        xfile["Sheet 1"]['B61'] = Area_std
        xfile["Sheet 1"]['B62'] = Area_gr_m
        xfile["Sheet 1"]['B63'] = Area_gr_max
        xfile["Sheet 1"]['B64'] = Area_gr_std

        xfile.save(path2 + model + '\Variables.xlsx')

        print("--- (8/14) Extracting Waveform Shape Variables took %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time_ws)/60), (time.time() - start_time_ws) % 60))
        print("--- (9/14) Finished Extracting Waveform Shape Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60), (time.time() - start_time) % 60))

        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        start_time_con = time.time()
        print("--- (10/14) Starting to Extract Waveform Shape Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60) , (time.time() - start_time) % 60))



        #v = vr N = nn and MEA = mea

        vr = pd.read_excel(path2 + model +'\Raster.xls').to_numpy()


        d2 = 18  # max connection distance squared

        Raster = np.zeros([nn, 10000])
        TE = np.zeros([nn, nn])

        for i in range(np.shape(vr)[0]):
            Raster[int(vr[i, 1]), int(vr[i, 0]/0.1)] = 1

        def row(A, i):
            return np.where(A == i)[0]  # returns row of neuron on MEA

        def col(A, i):
            return np.where(A == i)[1]  # Returns Column

        COL = []
        ROW = []

        for s in range(nn):
            COL = np.append(COL, col(mea, s))
            ROW = np.append(ROW, row(mea, s))

        freq = np.sum(Raster, axis=1);


        for i in range(nn):
            for j in range(nn):
                if i == j:
                    pass
                elif np.sqrt((ROW[i]-ROW[j])**2+(COL[i]-COL[j])**2)>np.sqrt(d2):
                    pass
                else:
                    TE[i, j] = nitime.algorithms.transfer_entropy(Raster[i, :], Raster[j, :], lag=3)



        Thr = 5.5*10**-5

        TEB = (Thr < TE)*1
        TEBa = (Thr < TE)*1
        TEBb = (Thr < TE)*1
        TEBg = (Thr < TE)*1

        p1 = 0
        p2 = 12.5
        p3 = 62.5
        p4 = 80.5

        for i in range(nn):
            for j in range(nn):
                if i == j:
                    pass
                elif ((freq[i] > p1 and freq[i] < p2) or (freq[j] > p1 and freq[j] < p2)):
                    TEBb[i,j] = 0
                    TEBg[i,j] = 0
                elif ((freq[i] > p2 and freq[i] < p3) or (freq[j] > p2 and freq[j] < p3)):
                    TEBa[i, j] = 0
                    TEBg[i, j] = 0
                elif ((freq[i] > p3 and freq[i] < p4) or (freq[j] > p3 and freq[j] < p4)):
                    TEBa[i, j] = 0
                    TEBb[i, j] = 0


        TEB1 = np.sum(TEB, axis=0)
        TEBa1 = np.sum(TEBa, axis=0)
        TEBb1 = np.sum(TEBb, axis=0)
        TEBg1 = np.sum(TEBg, axis=0)

        data1 = data*0
        data2 = data*0
        data3 = data*0
        data4 = data*0

        for i in range(np.shape(mea)[1] * np.shape(mea)[0]):
            data1[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = TEB1[i]
            data2[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = TEBa1[i]
            data3[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = TEBb1[i]
            data4[int(np.where(mea == i)[0]), int(np.where(mea == i)[1])] = TEBg1[i]


        Gr_x1, Gr_y1 = gradient(data1, ss)
        Gr_TEB1 = np.power(np.power(Gr_x1, 2) + np.power(Gr_y1, 2), (1 / 2))
        Gr_x2, Gr_y2 = gradient(data2, ss)
        Gr_TEBa1 = np.power(np.power(Gr_x2, 2) + np.power(Gr_y2, 2), (1 / 2))
        Gr_x3, Gr_y3 = gradient(data3, ss)
        Gr_TEBb1 = np.power(np.power(Gr_x3, 2) + np.power(Gr_y3, 2), (1 / 2))
        Gr_x4, Gr_y4 = gradient(data4, ss)
        Gr_TEBg1 = np.power(np.power(Gr_x4, 2) + np.power(Gr_y4, 2), (1 / 2))


        Con_m = np.mean(TEB1)
        Con_max = np.max(TEB1)
        Con_std = np.std(TEB1)
        Con_tot = np.sum(TEB1)
        Con_gr_m = np.mean(Gr_TEB1)
        Con_gr_max = np.max(Gr_TEB1)
        Con_gr_std = np.std(Gr_TEB1)

        Cona_m = np.mean(TEBa1)
        Cona_max = np.max(TEBa1)
        Cona_std = np.std(TEBa1)
        Cona_tot = np.sum(TEBa1)
        Cona_gr_m = np.mean(Gr_TEBa1)
        Cona_gr_max = np.max(Gr_TEBa1)
        Cona_gr_std = np.std(Gr_TEBa1)

        Conb_m = np.mean(TEBb1)
        Conb_max = np.max(TEBb1)
        Conb_std = np.std(TEBb1)
        Conb_tot = np.sum(TEBb1)
        Conb_gr_m = np.mean(Gr_TEBb1)
        Conb_gr_max = np.max(Gr_TEBb1)
        Conb_gr_std = np.std(Gr_TEBb1)

        Cong_m = np.mean(TEBg1)
        Cong_max = np.max(TEBg1)
        Cong_std = np.std(TEBg1)
        Cong_tot = np.sum(TEBg1)
        Cong_gr_m = np.mean(Gr_TEBg1)
        Cong_gr_max = np.max(Gr_TEBg1)
        Cong_gr_std = np.std(Gr_TEBg1)

        xfile = openpyxl.load_workbook(path2 + model + '\Variables.xlsx')

        xfile["Sheet 1"]['A65'] = 'MEAN NUMBER OF CONNECTIONS'
        xfile["Sheet 1"]['A66'] = 'MAX NUMBER OF CONNECTIONS'
        xfile["Sheet 1"]['A67'] = 'NUMBER OF CONNECTIONS STD'
        xfile["Sheet 1"]['A68'] = 'NUMBER OF CONNECTIONS TOTAL'
        xfile["Sheet 1"]['A69'] = 'MEAN NUMBER OF CONNECTIONS GRADIENT'
        xfile["Sheet 1"]['A70'] = 'MAX NUMBER OF CONNECTIONS GRADIENT'
        xfile["Sheet 1"]['A71'] = 'NUMBER OF CONNECTIONS GRADIENT STD'

        xfile["Sheet 1"]['B65'] = Con_m
        xfile["Sheet 1"]['B66'] = Con_max
        xfile["Sheet 1"]['B67'] = Con_std
        xfile["Sheet 1"]['B68'] = Con_tot
        xfile["Sheet 1"]['B69'] = Con_gr_m
        xfile["Sheet 1"]['B70'] = Con_gr_max
        xfile["Sheet 1"]['B71'] = Con_gr_std


        xfile["Sheet 1"]['A72'] = 'MEAN NUMBER OF CONNECTIONS IN THE ALPHA BAND'
        xfile["Sheet 1"]['A73'] = 'MAX NUMBER OF CONNECTIONS IN THE ALPHA BAND'
        xfile["Sheet 1"]['A74'] = 'NUMBER OF CONNECTIONS IN THE ALPHA BAND STD'
        xfile["Sheet 1"]['A75'] = 'NUMBER OF CONNECTIONS IN THE ALPHA BAND TOTAL'
        xfile["Sheet 1"]['A76'] = 'MEAN NUMBER OF CONNECTIONS IN THE ALPHA BAND GRADIENT'
        xfile["Sheet 1"]['A77'] = 'MAX NUMBER OF CONNECTIONS IN THE ALPHA BAND GRADIENT'
        xfile["Sheet 1"]['A78'] = 'NUMBER OF CONNECTIONS IN THE ALPHA BAND GRADIENT STD'

        xfile["Sheet 1"]['B72'] = Cona_m
        xfile["Sheet 1"]['B73'] = Cona_max
        xfile["Sheet 1"]['B74'] = Cona_std
        xfile["Sheet 1"]['B75'] = Cona_tot
        xfile["Sheet 1"]['B76'] = Cona_gr_m
        xfile["Sheet 1"]['B77'] = Cona_gr_max
        xfile["Sheet 1"]['B78'] = Cona_gr_std

        xfile["Sheet 1"]['A79'] = 'MEAN NUMBER OF CONNECTIONS IN THE BETA BAND'
        xfile["Sheet 1"]['A80'] = 'MAX NUMBER OF CONNECTIONS IN THE BETA BAND'
        xfile["Sheet 1"]['A81'] = 'NUMBER OF CONNECTIONS IN THE BETA STD'
        xfile["Sheet 1"]['A82'] = 'NUMBER OF CONNECTIONS IN THE BETA BAND TOTAL'
        xfile["Sheet 1"]['A83'] = 'MEAN NUMBER OF CONNECTIONS IN THE BETA BAND GRADIENT'
        xfile["Sheet 1"]['A84'] = 'MAX NUMBER OF CONNECTIONS IN THE BETA BAND GRADIENT'
        xfile["Sheet 1"]['A85'] = 'NUMBER OF CONNECTIONS IN THE BETA BAND GRADIENT STD'

        xfile["Sheet 1"]['B79'] = Conb_m
        xfile["Sheet 1"]['B80'] = Conb_max
        xfile["Sheet 1"]['B81'] = Conb_std
        xfile["Sheet 1"]['B82'] = Conb_tot
        xfile["Sheet 1"]['B83'] = Conb_gr_m
        xfile["Sheet 1"]['B84'] = Conb_gr_max
        xfile["Sheet 1"]['B85'] = Conb_gr_std

        xfile["Sheet 1"]['A86'] = 'MEAN NUMBER OF CONNECTIONS IN THE GAMMA BAND'
        xfile["Sheet 1"]['A87'] = 'MAX NUMBER OF CONNECTIONS IN THE GAMMA BAND'
        xfile["Sheet 1"]['A88'] = 'NUMBER OF CONNECTIONS IN THE GAMMA STD'
        xfile["Sheet 1"]['A89'] = 'NUMBER OF CONNECTIONS IN THE GAMMA BAND TOTAL'
        xfile["Sheet 1"]['A90'] = 'MEAN NUMBER OF CONNECTIONS IN THE GAMMA BAND GRADIENT'
        xfile["Sheet 1"]['A91'] = 'MAX NUMBER OF CONNECTIONS IN THE GAMMA BAND GRADIENT'
        xfile["Sheet 1"]['A92'] = 'NUMBER OF CONNECTIONS IN THE GAMMA BAND GRADIENT STD'

        xfile["Sheet 1"]['B86'] = Cong_m
        xfile["Sheet 1"]['B87'] = Cong_max
        xfile["Sheet 1"]['B88'] = Cong_std
        xfile["Sheet 1"]['B89'] = Cong_tot
        xfile["Sheet 1"]['B90'] = Cong_gr_m
        xfile["Sheet 1"]['B91'] = Cong_gr_max
        xfile["Sheet 1"]['B92'] = Cong_gr_std

        xfile.save(path2 + model + '\Variables.xlsx')

        print("--- (11/14) Extracting Connectivity Variables took %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time_con)/60), (time.time() - start_time_con) % 60))
        print("--- (12/14) Finished Extracting Connectivity Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60), (time.time() - start_time) % 60))
        print("--- (13/14) Finished Extracting All Variables at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60), (time.time() - start_time) % 60))
        print("--- (14/14) Done! at %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time)/60), (time.time() - start_time) % 60))

        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################
        ############################################################################################################################