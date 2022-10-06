#===============================
# AUTHOR: Gavin Kress NeuroDetect Inc.
# Date: 10/5/2022


import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # or any {'0', '1', '2'}
os.environ["PATH"] += os.pathsep + 'C:/Program Files/Graphviz/bin/'
import tensorflow as tf
import numpy as np
import openpyxl
import pandas as pd
import time
import scipy.stats as stats
import win32com.client as win32
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from sklearn.model_selection import cross_val_score
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import StratifiedKFold
from sklearn import tree
from keras.models import Sequential
from keras.layers import Dense
from keras import utils
from keras.wrappers.scikit_learn import KerasClassifier
import random
from matplotlib import pyplot as plt
from tqdm import tqdm
from joblib import dump, load
from keras_visualizer import visualizer
from sklearn import svm
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler
from sklearn import preprocessing
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from matplotlib import colors
import warnings
import lightgbm as lgb
warnings.filterwarnings("ignore")

def tensorflow_verbose_off():

    try:
        os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

        from tensorflow.python.util import deprecation

        tf.compat.v1.logging.set_verbosity(tf.compat.v1.logging.ERROR)

        def deprecated(date, instructions, warn_once=True):  # pylint: disable=unused-argument
            def deprecated_wrapper(func):
                return func
            return deprecated_wrapper

        deprecation.deprecated = deprecated

    except ImportError:
        pass


tensorflow_verbose_off()

start_time = time.time()

######################################################################################################################

path = 'C:/Users/gavin/PycharmProjects\Variable_Analysis\Models'
strArr = ['\Disease', '\Healthy', '\DiseaseLOF']
Num_models = 76
av = 0.02  # alpha value

m = -1
Vars = np.zeros([Num_models*len(strArr), 92])  # Obtain Variables

for ww in strArr:
    for kk in range(1, int(Num_models + 1)):
        m += 1
        model = ww + str(kk)
        Vars[m, :] = pd.read_excel(path + model + '\Variables.xlsx', header=None).to_numpy()[:, 1]



Vars_Disease = Vars[:Num_models, :]
Vars_Healthy = Vars[Num_models:2*Num_models, :]
Vars_DiseaseLOF = Vars[2*Num_models:, :]
names = pd.read_excel(path + model + '\Variables.xlsx', header=None)[0]
######################################################################################################################
print("--- Passing Variables Through Statistical Modeling Module ---")
P_values = np.zeros([3, 92])  # Obtain P Values

for k in range(92):
    P_values[:, k] = [stats.ttest_ind(Vars_Disease[:, k], Vars_Healthy[:, k], nan_policy='omit').pvalue,\
     stats.ttest_ind(Vars_DiseaseLOF[:, k], Vars_Healthy[:, k]).pvalue,\
     stats.ttest_ind(Vars_Disease[:, k], Vars_DiseaseLOF[:, k]).pvalue]

#####################################################################################################################

# Create P Value data frame and place in excel sheet, then format excel sheet
print("--- Creating Variable Data Frame ---")
P_values_df = pd.DataFrame(data = {'Variable': pd.read_excel(path + model + '\Variables.xlsx', header=None).to_numpy()[:, 0],\
                                   'Healthy/Disease': P_values[0, :], 'Healthy/DiseaseLOF': P_values[1, :],\
                                   'Disease/DiseaseLOF': P_values[2, :]})

P_values_df.to_excel(excel_writer=path + "/P_Values.xlsx")

excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(path + "/P_Values.xlsx")
ws = wb.Worksheets("Sheet1")
ws.Columns.AutoFit()
wb.Save()
excel.Application.Quit()

xfile = openpyxl.load_workbook(path + '/P_Values.xlsx')
xfiles = xfile.active
red_fill = PatternFill(bgColor="FFC7CE")
rule = CellIsRule(operator="lessThan", formula=[str(av)], fill=red_fill)
xfiles.conditional_formatting.add('B1:E93', rule)
redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
xfile.save(path + '/P_Values.xlsx')

##################################################################################################################

# Create Inputs for NN
NN_Input = np.array([Vars_Disease, Vars_Healthy, Vars_DiseaseLOF])
NN_Input_SS = np.zeros([92])  # Will become index list of SS Variables that go into NN

for m in range(len(strArr)):
    for j in range(92):
        if any(P_values[:, j] < av):
            NN_Input_SS[j] = 1
        else:
            NN_Input[:, :, j] = [[float("NaN")]*Num_models, [float("NaN")]*Num_models, [float("NaN")]*Num_models]

NN_Input_SS = np.array(np.where(NN_Input_SS == 1)[0])
names = names[NN_Input_SS]

###############################################################################################################################################

# Flatten NN variables into df
NN_Input_df = np.empty([len(strArr)*Num_models, len(NN_Input_SS) + 1])
NN_Input_df[0:Num_models, 0:len(NN_Input_SS)] = Vars_Disease[:, NN_Input_SS]
NN_Input_df[Num_models:2*Num_models, 0:len(NN_Input_SS)] = Vars_Healthy[:, NN_Input_SS]
NN_Input_df[2*Num_models:, 0:len(NN_Input_SS)] = Vars_DiseaseLOF[:, NN_Input_SS]
NN_Input_df = pd.DataFrame(NN_Input_df)

NN_Input_df[len(NN_Input_SS)][0:Num_models] = ['Disease']*Num_models
NN_Input_df[len(NN_Input_SS)][Num_models:2*Num_models] = ['Healthy']*Num_models
NN_Input_df[len(NN_Input_SS)][2*Num_models:] = ['DiseaseLOF']*Num_models


# split into input (X) and output (Y) variables
X = NN_Input_df.T[0:len(NN_Input_SS)].T.astype(float)
Y = NN_Input_df[len(NN_Input_SS)]

# encode class values as integers
encoder = LabelEncoder()
encoder.fit(Y)
encoded_Y = encoder.transform(Y)

# Normalize input Data

def normalize_input_data(x, axes=[0, 1], epsilon=1e-8):
    mean, variance = tf.nn.moments(x, axes=axes)
    x_normed = (x - mean) / tf.sqrt(variance + epsilon) # epsilon to avoid dividing by zero
    return x_normed

# create and test estimator
epo = 800
bs = 80


def create_model():
    model = Sequential()  # create model
    model.add(Dense(len(NN_Input_SS), input_dim=len(NN_Input_SS), activation='relu'))
    model.add(Dense(round(len(NN_Input_SS)), activation='relu'))
    model.add(Dense(round(len(NN_Input_SS)), activation='relu'))
    model.add(Dense(round(len(NN_Input_SS)), activation='relu'))
    model.add(Dense(3, activation='softmax'))
    model.compile(loss='sparse_categorical_crossentropy', optimizer='adam', metrics=['accuracy'])  # Compile model
    return model

def kfoldtesting_NN(X, verbose: int):
    X = pd.DataFrame(normalize_input_data(tf.convert_to_tensor(X)).numpy())
    print("--- Training Neural Network ---")
    ModelClassifier = KerasClassifier(build_fn=create_model, epochs=epo, batch_size=bs, verbose=verbose)
    print("--- Testing Neural Network ---")
    kfold = StratifiedKFold(n_splits=10)
    results = cross_val_score(ModelClassifier, X, encoded_Y, cv=kfold)
    print("--- Results: %.2f%% (STD %.2f%%) (MAX %.2f%%)---" % (results.mean()*100, results.std()*100, results.max()))


def Train_Serial_NN(Training_Models, X, verbose: str, threshold: float):

    X = pd.DataFrame(normalize_input_data(tf.convert_to_tensor(X)).numpy())

    # fit and serialize tested model
    TM = Training_Models # Training Models
    ModelFit = create_model()
    DisTrain = random.sample(range(0, Num_models), TM)
    HealTrain = random.sample(range(Num_models, 2*Num_models), TM)
    DisLOFTrain = random.sample(range(2*Num_models, 3*Num_models), TM)
    Train = np.ndarray.flatten(np.array([DisTrain, HealTrain, DisLOFTrain]))  # Training Data

    def find_missing(lst):
        return [x for x in range(0, len(strArr)*Num_models) if x not in lst]

    Test = find_missing(Train)  # Testing Data

    ModelFit.fit(np.array(X)[Train, :], np.array(encoded_Y)[Train], epochs=epo, batch_size=bs, verbose=0)  # Train Model

    Test_Results = np.argmax(ModelFit(pd.DataFrame.to_numpy(X)[Test, :]),1)
    Performance = (encoded_Y[Test] == Test_Results).sum()/len(Test_Results)*100

    nn = 0
    D_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    D_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    D_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    D_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    D_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    D_pre = D_TP/(D_TP+D_FP)
    D_rec = D_TP/(D_TP+D_FN)
    D_F1 = 2*D_pre*D_rec/(D_pre+D_rec)


    nn = 2
    H_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    H_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    H_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    H_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    H_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    HP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    H_pre = H_TP/(H_TP+H_FP)
    H_rec = H_TP/(H_TP+H_FN)
    H_F1 = 2*H_pre*H_rec/(H_pre+H_rec)

    nn = 1
    DL_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    DL_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    DL_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    DL_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    DL_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DLP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    DL_pre = DL_TP/(DL_TP+DL_FP)
    DL_rec = DL_TP/(DL_TP+DL_FN)
    DL_F1 = 2*DL_pre*DL_rec/(DL_pre+DL_rec)


    if Performance > threshold*100:
        ModelFit.save(path + "/NN_Serial" + "/NN_model_" + str(round(Performance, 2)) + "_Overall_"\
                      + str(round(DP, 2))+ "_"+str(round(D_TPR, 2))+ "_"+str(round(D_FPR, 2)) + "_"+str(round(D_pre, 2))+ "_"+str(round(D_rec, 2))+ "_"+str(round(D_F1, 2))+ "_GOF_"\
                      + str(round(HP, 2))+ "_"+str(round(H_TPR, 2))+ "_"+str(round(H_FPR, 2)) + "_"+str(round(H_pre, 2))+ "_"+str(round(H_rec, 2))+ "_"+str(round(H_F1, 2)) + "_Healthy_"\
                      + str(round(DLP, 2))+ "_"+str(round(DL_TPR, 2))+ "_"+str(round(DL_FPR, 2)) + "_"+str(round(DL_pre, 2))+ "_"+str(round(DL_rec, 2))+ "_"+str(round(DL_F1, 2))+  "_LOF_"\
                      + ".h5")
        file = path + "/NN_Serial" + "/NN_model_" + str(round(Performance, 2)) + "_Overall_"\
                      + str(round(DP, 2))+ "_"+str(round(D_TPR, 2))+ "_"+str(round(D_FPR, 2)) + "_"+str(round(D_pre, 2))+ "_"+str(round(D_rec, 2))+ "_"+str(round(D_F1, 2))+ "_GOF_"\
                      + str(round(HP, 2))+ "_"+str(round(H_TPR, 2))+ "_"+str(round(H_FPR, 2)) + "_"+str(round(H_pre, 2))+ "_"+str(round(H_rec, 2))+ "_"+str(round(H_F1, 2)) + "_Healthy_"\
                      + str(round(DLP, 2))+ "_"+str(round(DL_TPR, 2))+ "_"+str(round(DL_FPR, 2)) + "_"+str(round(DL_pre, 2))+ "_"+str(round(DL_rec, 2))+ "_"+str(round(DL_F1, 2))+  "_LOF_"
        visualizer(ModelFit, format='png', view=False, filename=file)

    if verbose == 'on':
        print("---  Model Got %.2f Overall Percent Correct (%.2f GOF, %.2f Healthy, %.2f LOF) ---" % (Performance, DP, HP, DLP))
        print("---  Finished after %.0f minutes and %.2f seconds ---" % (np.floor((time.time() - start_time) / 60), (time.time() - start_time) % 60))
    elif verbose == 'model':
        print("--- Model Got %.2f Overall Percent Correct (%.2f GOF, %.2f Healthy, %.2f LOF) ---" % (Performance, DP, HP, DLP))



def Create_and_fit_Decision_Tree(Training_Models, X, plot: str, threshold: float, verbose: str):
    if verbose == 'on': print("--- Creating DecisionTree Classifier ---")
    TM = Training_Models  # Training Models
    DisTrain = random.sample(range(0, Num_models), TM)
    HealTrain = random.sample(range(Num_models, 2 * Num_models), TM)
    DisLOFTrain = random.sample(range(2 * Num_models, 3 * Num_models), TM)
    Train = np.ndarray.flatten(np.array([DisTrain, HealTrain, DisLOFTrain]))  # Training Data

    def find_missing(lst):
        return [x for x in range(0, len(strArr) * Num_models) if x not in lst]

    Test = find_missing(Train)  # Testing Data

    model_tree = tree.DecisionTreeClassifier()
    Tree = model_tree.fit(np.array(X)[Train, :], np.array(encoded_Y)[Train])
    Predicted = Tree.predict(np.array(X)[Test, :])
    Actual = np.array(encoded_Y)[Test]
    Accuracy = np.sum((Predicted==Actual)*1.)/len(Predicted)
    Accuracy_Healthy = np.sum((Predicted[np.where(Actual == 0)] == Actual[np.where(Actual == 0)]) * 1.) / len(
        Actual[np.where(Actual == 0)])
    Accuracy_DiseaseLOF = np.sum((Predicted[np.where(Actual == 1)] == Actual[np.where(Actual == 1)]) * 1.) / len(
        Actual[np.where(Actual == 1)])
    Accuracy_Disease = np.sum((Predicted[np.where(Actual == 2)] == Actual[np.where(Actual == 2)]) * 1.) / len(
        Actual[np.where(Actual == 2)])


    Test_Results = Predicted
    nn = 0
    D_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    D_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    D_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    D_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    D_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    D_pre = D_TP/(D_TP+D_FP)
    D_rec = D_TP/(D_TP+D_FN)
    D_F1 = 2*D_pre*D_rec/(D_pre+D_rec)


    nn = 2
    H_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    H_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    H_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    H_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    H_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    HP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    H_pre = H_TP/(H_TP+H_FP)
    H_rec = H_TP/(H_TP+H_FN)
    H_F1 = 2*H_pre*H_rec/(H_pre+H_rec)

    nn = 1
    DL_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    DL_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    DL_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    DL_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    DL_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DLP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    DL_pre = DL_TP/(DL_TP+DL_FP)
    DL_rec = DL_TP/(DL_TP+DL_FN)
    DL_F1 = 2*DL_pre*DL_rec/(DL_pre+DL_rec)



    if verbose == 'on': print("--- Decision Tree Model Got %.2f Overall Percent Correct (%.2f GOF, %.2f Healthy, %.2f LOF) ---" % (
    Accuracy, Accuracy_Disease, Accuracy_Healthy, Accuracy_DiseaseLOF))

    if plot == 'save' and Accuracy>threshold:
        import graphviz
        dot_data = tree.export_graphviz(Tree, out_file=None, filled=True, special_characters=True, feature_names=np.array(names))
        graph = graphviz.Source(dot_data)
        graph.render(path + '/Decision Tree Serial/DT_model_'+ str(round(Accuracy*100, 2)) + "_Overall_" \
                     + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
                     + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
                     + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_image")

    else:
        pass
    if Accuracy > threshold:
        dump(Tree, path + '/Decision Tree Serial/' + "/DT_model_" + str(round(Accuracy*100, 2)) + "_Overall_" \
             + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
             + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
             + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_" \
             + ".joblib")
    return Tree


def Grid_Search_SVM(Training_Models, X):
    from sklearn.utils.fixes import loguniform
    from sklearn.model_selection import GridSearchCV
    Training_Models = 50
    TM = Training_Models  # Training Models
    DisTrain = random.sample(range(0, Num_models), TM)
    HealTrain = random.sample(range(Num_models, 2 * Num_models), TM)
    DisLOFTrain = random.sample(range(2 * Num_models, 3 * Num_models), TM)
    Train = np.ndarray.flatten(np.array([DisTrain, HealTrain, DisLOFTrain]))  # Training Data

    def find_missing(lst):
        return [x for x in range(0, len(strArr) * Num_models) if x not in lst]

    Test = find_missing(Train)  # Testing Data

    param_grid = {'C': np.arange(0,1,0.001), 'gamma': [0.013], 'kernel': ['rbf'], 'class_weight': ['balanced']}
    clf = GridSearchCV(svm.SVC(cache_size=1000), param_grid, verbose=10)
    clf.fit(np.array(X)[Train, :], np.array(encoded_Y)[Train])
    clf.cv_results_['params'][clf.best_index_]
    clf.best_score_

def Create_and_fit_Support_Vector_Machine(Training_Models, X, plot: str, threshold: float, verbose: str, scale: bool):
    if verbose == 'on': print("--- Creating SVM Classifier ---")
    TM = Training_Models  # Training Models
    DisTrain = random.sample(range(0, Num_models), TM)
    HealTrain = random.sample(range(Num_models, 2 * Num_models), TM)
    DisLOFTrain = random.sample(range(2 * Num_models, 3 * Num_models), TM)
    Train = np.ndarray.flatten(np.array([DisTrain, HealTrain, DisLOFTrain]))  # Training Data

    def find_missing(lst):
        return [x for x in range(0, len(strArr) * Num_models) if x not in lst]

    Test = find_missing(Train)  # Testing Data

    model_SVM = svm.SVC(cache_size=4000, C = 0.4)

    # Preprocess Data
    if scale == True:
        scaler = preprocessing.StandardScaler().fit(X)
        X_scaled = scaler.transform(X)
        X = X_scaled
    # Train Data

    SVM = model_SVM.fit(np.array(X)[Train, :], np.array(encoded_Y)[Train])
    Predicted = SVM.predict(np.array(X)[Test, :])
    Actual = np.array(encoded_Y)[Test]
    Accuracy = np.sum((Predicted==Actual)*1.)/len(Predicted)
    Accuracy_Healthy = np.sum((Predicted[np.where(Actual == 0)] == Actual[np.where(Actual == 0)]) * 1.) / len(
        Actual[np.where(Actual == 0)])
    Accuracy_DiseaseLOF = np.sum((Predicted[np.where(Actual == 1)] == Actual[np.where(Actual == 1)]) * 1.) / len(
        Actual[np.where(Actual == 1)])
    Accuracy_Disease = np.sum((Predicted[np.where(Actual == 2)] == Actual[np.where(Actual == 2)]) * 1.) / len(
        Actual[np.where(Actual == 2)])


    Test_Results = Predicted
    nn = 0
    D_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    D_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    D_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    D_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    D_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    D_pre = D_TP/(D_TP+D_FP)
    D_rec = D_TP/(D_TP+D_FN)
    D_F1 = 2*D_pre*D_rec/(D_pre+D_rec)


    nn = 2
    H_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    H_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    H_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    H_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    H_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    HP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    H_pre = H_TP/(H_TP+H_FP)
    H_rec = H_TP/(H_TP+H_FN)
    H_F1 = 2*H_pre*H_rec/(H_pre+H_rec)

    nn = 1
    DL_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    DL_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    DL_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    DL_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    DL_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DLP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    DL_pre = DL_TP/(DL_TP+DL_FP)
    DL_rec = DL_TP/(DL_TP+DL_FN)
    DL_F1 = 2*DL_pre*DL_rec/(DL_pre+DL_rec)




    if verbose == 'on': print("--- SVM Model Got %.2f Overall Percent Correct (%.2f GOF, %.2f Healthy, %.2f LOF) ---" % (
    Accuracy, Accuracy_Disease, Accuracy_Healthy, Accuracy_DiseaseLOF))

    if plot == 'save' and Accuracy>threshold:
        h = .2  # step size in the mesh
        XX = np.array(X)
        ww = np.array([[0, 1], [33, 3], [34, 60], [63, 61]])
        plt.ioff()
        fig = plt.figure(figsize=(14,14),  dpi = 200)
        plt.ioff()
        for i in range(4):
            plt.subplot(2, 2, i + 1)
            plt.subplots_adjust(wspace=0.4, hspace=0.4)

            # Variables to plot
            v = ww[i, :]
            #v = np.random.randint(0,[65,65])

            # create a mesh to plot in
            x_min, x_max = XX[:, v[0]].min() - 1, XX[:, v[0]].max() + 1
            y_min, y_max = XX[:, v[1]].min() - 1, XX[:, v[1]].max() + 1
            xx, yy = np.meshgrid(np.arange(x_min, x_max, h),
                                 np.arange(y_min, y_max, h))

            w = np.c_[np.repeat(np.c_[xx.ravel(), yy.ravel()], 32, axis = 1), xx.ravel()]


            Z = SVM.predict(w)
            Z = Z.reshape(xx.shape)
            plt.contourf(xx, yy, Z, cmap=plt.cm.coolwarm, alpha=0.8)

            y = np.array(encoded_Y)
            plt.scatter(XX[:, v[0]], XX[:, v[1]], c=y, cmap=plt.cm.coolwarm)
            plt.xlabel(np.array(names)[v[0]])
            plt.ylabel(np.array(names)[v[1]])
            plt.xlim(xx.min(), xx.max())
            plt.ylim(yy.min(), yy.max())
            plt.xticks(())
            plt.yticks(())

        file = path + '/SVM Serial/SVM_model_' + str(round(Accuracy * 100, 2)) + "_Overall_" \
               + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
               + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
               + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_image.png"

        plt.savefig(file)
    else:
        pass

    if Accuracy > threshold:
        dump(SVM, path + '/SVM Serial/' + "/SVM_model_" + str(round(Accuracy*100, 2)) + "_Overall_" \
             + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
             + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
             + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_" \
             + ".joblib")
    return SVM



def Create_and_fit_Naive_Bays(Training_Models, X, plot: str, threshold: float, verbose: str, type: str):
    if verbose == 'on': print("--- Creating Gaussian Naive Bayes Classifier ---")
    TM = Training_Models  # Training Models
    DisTrain = random.sample(range(0, Num_models), TM)
    HealTrain = random.sample(range(Num_models, 2 * Num_models), TM)
    DisLOFTrain = random.sample(range(2 * Num_models, 3 * Num_models), TM)
    Train = np.ndarray.flatten(np.array([DisTrain, HealTrain, DisLOFTrain]))  # Training Data

    def find_missing(lst):
        return [x for x in range(0, len(strArr) * Num_models) if x not in lst]

    Test = find_missing(Train)  # Testing Data

    if type == 'Gaussian':
        gnb = GaussianNB()
    elif type == 'Catagorical':
        from sklearn.naive_bayes import CategoricalNB
        gnb = CategoricalNB()
    elif type == 'Multinomial':
        from sklearn.naive_bayes import MultinomialNB
        gnb = MultinomialNB()
    elif type == 'Complement':
        from sklearn.naive_bayes import ComplementNB
        gnb = ComplementNB()
    elif type == 'Bernoulli':
        from sklearn.naive_bayes import BernoulliNB
        gnb = BernoulliNB()

    GNB = gnb.fit(np.array(X)[Train, :], np.array(encoded_Y)[Train])
    Predicted = GNB.predict(np.array(X)[Test, :])
    Actual = np.array(encoded_Y)[Test]
    Accuracy = np.sum((Predicted==Actual)*1.)/len(Predicted)
    Accuracy_Healthy = np.sum((Predicted[np.where(Actual == 0)] == Actual[np.where(Actual == 0)]) * 1.) / len(
        Actual[np.where(Actual == 0)])
    Accuracy_DiseaseLOF = np.sum((Predicted[np.where(Actual == 1)] == Actual[np.where(Actual == 1)]) * 1.) / len(
        Actual[np.where(Actual == 1)])
    Accuracy_Disease = np.sum((Predicted[np.where(Actual == 2)] == Actual[np.where(Actual == 2)]) * 1.) / len(
        Actual[np.where(Actual == 2)])


    Test_Results = Predicted
    nn = 0
    D_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    D_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    D_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    D_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    D_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    D_pre = D_TP/(D_TP+D_FP)
    D_rec = D_TP/(D_TP+D_FN)
    D_F1 = 2*D_pre*D_rec/(D_pre+D_rec)


    nn = 2
    H_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    H_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    H_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    H_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    H_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    HP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    H_pre = H_TP/(H_TP+H_FP)
    H_rec = H_TP/(H_TP+H_FN)
    H_F1 = 2*H_pre*H_rec/(H_pre+H_rec)

    nn = 1
    DL_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    DL_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    DL_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    DL_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    DL_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DLP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    DL_pre = DL_TP/(DL_TP+DL_FP)
    DL_rec = DL_TP/(DL_TP+DL_FN)
    DL_F1 = 2*DL_pre*DL_rec/(DL_pre+DL_rec)





    if verbose == 'on': print("--- Gaussian Naive Bayes Got %.2f Overall Percent Correct (%.2f GOF, %.2f Healthy, %.2f LOF) ---" % (
    Accuracy, Accuracy_Disease, Accuracy_Healthy, Accuracy_DiseaseLOF))

    if plot == 'save' and Accuracy > threshold:
        X = np.array(X)
        h = 0.4
        Varis = np.array([[24, 53], [58, 21], [57, 42], [44, 1]])

        plt.ioff()
        fig = plt.figure(figsize=(14, 14), dpi=200)

        plt.ioff()
        for i in range(4):
            plt.subplot(2, 2, i + 1)
            v = Varis[i, :]

            # create a mesh to plot in
            x_min, x_max = X[:, v[0]].min() - 1, X[:, v[0]].max() + 1
            y_min, y_max = X[:, v[1]].min() - 1, X[:, v[1]].max() + 1
            xx, yy = np.meshgrid(np.arange(x_min, x_max, h),
                                 np.arange(y_min, y_max, h))



            w = np.c_[np.repeat(np.c_[xx.ravel(), yy.ravel()], 32, axis = 1), xx.ravel()]


            Z = GNB.predict(w)
            Z = Z.reshape(xx.shape)
            plt.contourf(xx, yy, Z, cmap=plt.cm.coolwarm, alpha=0.8)

            y = np.array(encoded_Y)
            plt.scatter(X[:, v[0]], X[:, v[1]], c=y, cmap=plt.cm.coolwarm )
            plt.xlabel(np.array(names)[v[0]])
            plt.ylabel(np.array(names)[v[1]])
            plt.xlim(xx.min(), xx.max())
            plt.ylim(yy.min(), yy.max())
            plt.xlabel(np.array(names)[v[0]])
            plt.ylabel(np.array(names)[v[1]])
        file = path + '/The GNB Serial/GNB_model_'+ str(round(Accuracy*100, 2)) + "_Overall_" \
               + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
               + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
               + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_image.png"
        plt.savefig(file)


    else:
        pass
    if Accuracy > threshold:
        dump(GNB, path + '/The GNB Serial/' + "/GNB_model_" + str(round(Accuracy*100, 2)) + "_Overall_" \
             + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
             + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
             + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_" \
             + ".joblib")
    return GNB




def Create_and_fit_LightGBM(Training_Models, X, threshold: float):
    TM = Training_Models  # Training Models
    DisTrain = random.sample(range(0, Num_models), TM)
    HealTrain = random.sample(range(Num_models, 2 * Num_models), TM)
    DisLOFTrain = random.sample(range(2 * Num_models, 3 * Num_models), TM)
    Train = np.ndarray.flatten(np.array([DisTrain, HealTrain, DisLOFTrain]))  # Training Data

    def find_missing(lst):
        return [x for x in range(0, len(strArr) * Num_models) if x not in lst]

    Test = find_missing(Train)  # Testing Data
    train_data = lgb.Dataset(np.array(X)[Train, :], label=np.array(encoded_Y)[Train], feature_name=list(names))
    params = {
              'boosting_type': 'gbdt',
              'objective': 'multiclass',
              'num_class': 3,
              'verbosity': -1}

    bst = lgb.train(params, train_data)
    Predicted = bst.predict(np.array(X)[Test, :])
    Predicted = np.argmax(Predicted,1)
    Actual = np.array(encoded_Y)[Test]
    Accuracy = np.sum((Predicted==Actual)*1.)/len(Predicted)
    Accuracy_Healthy = np.sum((Predicted[np.where(Actual == 0)] == Actual[np.where(Actual == 0)]) * 1.) / len(Actual[np.where(Actual == 0)])
    Accuracy_DiseaseLOF = np.sum((Predicted[np.where(Actual == 1)] == Actual[np.where(Actual == 1)]) * 1.) / len(Actual[np.where(Actual == 1)])
    Accuracy_Disease = np.sum((Predicted[np.where(Actual == 2)] == Actual[np.where(Actual == 2)]) * 1.) / len(Actual[np.where(Actual == 2)])


    Test_Results = Predicted
    nn = 0
    D_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    D_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    D_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    D_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    D_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    D_pre = D_TP/(D_TP+D_FP)
    D_rec = D_TP/(D_TP+D_FN)
    D_F1 = 2*D_pre*D_rec/(D_pre+D_rec)


    nn = 2
    H_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    H_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    H_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    H_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    H_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    HP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    H_pre = H_TP/(H_TP+H_FP)
    H_rec = H_TP/(H_TP+H_FN)
    H_F1 = 2*H_pre*H_rec/(H_pre+H_rec)

    nn = 1
    DL_TP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()
    DL_FP = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()
    DL_TPR = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()
    DL_FPR = ((encoded_Y[Test] != Test_Results) & (Test_Results == nn)).sum()/(encoded_Y[Test] != nn).sum()
    DL_FN = ((encoded_Y[Test] != Test_Results) & (encoded_Y[Test] == nn)).sum()
    DLP = ((encoded_Y[Test] == Test_Results) & (encoded_Y[Test] == nn)).sum()/(encoded_Y[Test] == nn).sum()*100
    DL_pre = DL_TP/(DL_TP+DL_FP)
    DL_rec = DL_TP/(DL_TP+DL_FN)
    DL_F1 = 2*DL_pre*DL_rec/(DL_pre+DL_rec)




    if Accuracy > threshold:
        bst.save_model(path + '/The GBDT Serial/' + "/GBT_model_" + str(round(Accuracy*100, 2)) + "_Overall_" \
                       + str(round(DP, 2)) + "_" + str(round(D_TPR, 2)) + "_" + str(round(D_FPR, 2)) + "_" + str(
            round(D_pre, 2)) + "_" + str(round(D_rec, 2)) + "_" + str(round(D_F1, 2)) + "_GOF_" \
                       + str(round(HP, 2)) + "_" + str(round(H_TPR, 2)) + "_" + str(round(H_FPR, 2)) + "_" + str(
            round(H_pre, 2)) + "_" + str(round(H_rec, 2)) + "_" + str(round(H_F1, 2)) + "_Healthy_" \
                       + str(round(DLP, 2)) + "_" + str(round(DL_TPR, 2)) + "_" + str(round(DL_FPR, 2)) + "_" + str(
            round(DL_pre, 2)) + "_" + str(round(DL_rec, 2)) + "_" + str(round(DL_F1, 2)) + "_LOF_" \
                       + ".txt")
    return bst

#######################################################################################################################################
######################################################################################################################################



for i in tqdm(range(100)):
     Train_Serial_NN(50, X, verbose='off', threshold=0)

for i in tqdm(range(100)):
    Tree = Create_and_fit_Decision_Tree(50, X, plot='save', threshold = 0, verbose='off')

for i in tqdm(range(100)):
    SVM = Create_and_fit_Support_Vector_Machine(50, X, plot='save', threshold = 0, verbose='off', scale=False)

for i in tqdm(range(100)):
    GNB = Create_and_fit_Naive_Bays(50, X, plot='save', threshold = 0, verbose='off', type = 'Gaussian')

for i in tqdm(range(100)):
    GBDT = Create_and_fit_LightGBM(50, X, threshold=0)
